from tkinter import *
from tkinter import filedialog
import openpyxl as xl
from openpyxl.chart import LineChart, Reference
import os
import sys
import PySimpleGUI as sg
from docx.shared import Cm
from docx.shared import Cm, Inches, Mm, Emu
from docxtpl import DocxTemplate, InlineImage

import random
import datetime
import plotly.graph_objects as go

def openFile():
    filepath = filedialog.askopenfilename()
    workbook = xl.load_workbook(filepath, data_only=True)
    sheet_1 = workbook['OCP Cluster - Scoring']
    categoryGroup = sheet_1.cell(8, 2).value
    print(categoryGroup)
    categoryGroup = sheet_1.cell(8, 2).value

    print(categoryGroup)
    categories = categoryGroup.split(", ")

    template = DocxTemplate(
        'Template - IBM - RedHat OCP Assessment Report v1.docx')
    table_cat1 = []
    table_cat2 = []
    table_cat3 = []
    table_cat4 = []
    table_cat5 = []
    max_score = [0, 0, 0, 0, 0]
    score = [0, 0, 0, 0, 0]
    for row in range(10, sheet_1.max_row + 1):
        if (sheet_1.cell(row, 1).value == categories[0]):
            max_score[0] += int(sheet_1.cell(row, 6).value*10)
            score[0] += int(sheet_1.cell(row, 6).value *
                            sheet_1.cell(row, 7).value)
            table_cat1.append({
                # 'Index': row-1,
                'Category': sheet_1.cell(row, 1).value,
                'Question': sheet_1.cell(row, 2).value,
                'Answer': sheet_1.cell(row, 3).value,
                'Comment': sheet_1.cell(row, 4).value
            })

        elif (sheet_1.cell(row, 1).value == categories[1]):
            max_score[1] += int(sheet_1.cell(row, 6).value*10)
            score[1] += int(sheet_1.cell(row, 6).value *
                            sheet_1.cell(row, 7).value)
            table_cat2.append({
                # 'Index': row-1,
                'Category': sheet_1.cell(row, 1).value,
                'Question': sheet_1.cell(row, 2).value,
                'Answer': sheet_1.cell(row, 3).value,
                'Comment': sheet_1.cell(row, 4).value
            })
        elif (sheet_1.cell(row, 1).value == categories[2]):
            max_score[2] += int(sheet_1.cell(row, 6).value*10)
            score[2] += int(sheet_1.cell(row, 6).value *
                            sheet_1.cell(row, 7).value)
            table_cat3.append({
                # 'Index': row-1,
                'Category': sheet_1.cell(row, 1).value,
                'Question': sheet_1.cell(row, 2).value,
                'Answer': sheet_1.cell(row, 3).value,
                'Comment': sheet_1.cell(row, 4).value
            })
        elif (sheet_1.cell(row, 1).value == categories[3]):
            max_score[3] += int(sheet_1.cell(row, 6).value*10)
            score[3] += int(sheet_1.cell(row, 6).value *
                            sheet_1.cell(row, 7).value)
            table_cat4.append({
                # 'Index': row-1,
                'Category': sheet_1.cell(row, 1).value,
                'Question': sheet_1.cell(row, 2).value,
                'Answer': sheet_1.cell(row, 3).value,
                'Comment': sheet_1.cell(row, 4).value
            })
        elif (sheet_1.cell(row, 1).value == categories[4]):
            max_score[4] += int(sheet_1.cell(row, 6).value*10)
            score[4] += int(sheet_1.cell(row, 6).value *
                            sheet_1.cell(row, 7).value)
            table_cat5.append({
                # 'Index': row-1,
                'Category': sheet_1.cell(row, 1).value,
                'Question': sheet_1.cell(row, 2).value,
                'Answer': sheet_1.cell(row, 3).value,
                'Comment': sheet_1.cell(row, 4).value
            })

    result = [100*(score[0]/max_score[0]), 100*(score[1]/max_score[1]), 100*(score[2]/max_score[2]),
              100*(score[3]/max_score[3]), 100*(score[4]/max_score[4])]
    print(score)
    print(max_score)
    print(result)

    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=result,
        theta=categories,
        fill='toself',
        name='OCP'
    ))

    fig.update_layout(
        polar=dict(
            radialaxis=dict(
                visible=True,
                range=[0, 100]
            )),
        showlegend=False
    )

    fig.write_image("general_graph.png")
    image = InlineImage(template, 'general_graph.png', Cm(10))

    context = {
        'client': 'Automated Report',
        'startmonth': 'May',
        'endmonth': 'June',
        'year': '2050',
        'table_cat1': table_cat1,
        'table_cat2': table_cat2,
        'table_cat3': table_cat3,
        'table_cat4': table_cat4,
        'table_cat5': table_cat5,
        'image': image
    }
    template.render(context)
    template.save('Automated_report.docx')
    exit(0)


window = Tk()
button = Button(text="Open", command=openFile)
button.pack()
window.mainloop()
